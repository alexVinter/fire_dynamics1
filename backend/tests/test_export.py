"""Tests for XLSX export endpoint."""
import json
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.quote import Quote, QuoteItem
from app.models.quote_result_line import QuoteResultLine
from app.models.sku import SKU
from app.models.technique import Technique
from app.models.user import User
from app.services.xlsx_export import _safe

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _token(client: TestClient, login: str, password: str) -> str:
    return client.post("/auth/login", json={"login": login, "password": password}).json()["access_token"]


def test_export_returns_xlsx(client: TestClient, manager_user: User, db: Session):
    tech = Technique(manufacturer="X", model="Y")
    db.add(tech)
    db.flush()

    sku = SKU(code="EX-001", name="Трубка", unit="шт")
    db.add(sku)
    db.flush()

    q = Quote(created_by=manager_user.id, status="calculated", zones_json=json.dumps([]))
    db.add(q)
    db.flush()
    db.add(QuoteItem(quote_id=q.id, technique_id=tech.id, qty=1))
    db.add(QuoteResultLine(quote_id=q.id, sku_id=sku.id, qty=5))
    db.commit()

    token = _token(client, "manager", "mgr123")
    resp = client.post(
        f"/quotes/{q.id}/export/xlsx",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MIME
    assert len(resp.content) > 0
    assert resp.content[:2] == b"PK"


def test_safe_prefixes_injection_chars():
    assert _safe("=cmd|'/C calc'!A0") == "'=cmd|'/C calc'!A0"
    assert _safe("+1+1") == "'+1+1"
    assert _safe("-1-1") == "'-1-1"
    assert _safe("@SUM(A1)") == "'@SUM(A1)"
    assert _safe("normal text") == "normal text"
    assert _safe("") == ""


def test_export_sanitizes_injection(client: TestClient, manager_user: User, db: Session):
    tech = Technique(manufacturer="X", model="Y")
    db.add(tech)
    db.flush()

    sku = SKU(code="=EVIL", name="+danger", unit="шт")
    db.add(sku)
    db.flush()

    q = Quote(created_by=manager_user.id, status="calculated", zones_json=json.dumps([]))
    db.add(q)
    db.flush()
    db.add(QuoteItem(quote_id=q.id, technique_id=tech.id, qty=1))
    db.add(QuoteResultLine(quote_id=q.id, sku_id=sku.id, qty=3, note="-note"))
    db.commit()

    token = _token(client, "manager", "mgr123")
    resp = client.post(
        f"/quotes/{q.id}/export/xlsx",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 200

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    data_row = ws[6]
    assert data_row[0].value == "'=EVIL"
    assert data_row[1].value == "'+danger"
    assert data_row[4].value == "'-note"


def test_export_no_results_returns_409(client: TestClient, manager_user: User, db: Session):
    tech = Technique(manufacturer="X", model="Y")
    db.add(tech)
    db.flush()

    q = Quote(created_by=manager_user.id, status="draft", zones_json=json.dumps([]))
    db.add(q)
    db.flush()
    db.add(QuoteItem(quote_id=q.id, technique_id=tech.id, qty=1))
    db.commit()

    token = _token(client, "manager", "mgr123")
    resp = client.post(
        f"/quotes/{q.id}/export/xlsx",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert resp.status_code == 409
