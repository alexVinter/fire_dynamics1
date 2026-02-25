from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.quote import Quote
from app.models.quote_result_line import QuoteResultLine
from app.models.sku import SKU
from app.models.user import User


def xlsx_export(db: Session, quote_id: int) -> bytes:
    quote = db.get(Quote, quote_id)
    if quote is None:
        raise ValueError(f"Quote {quote_id} not found")

    creator = db.get(User, quote.created_by)
    manager_name = creator.login if creator else str(quote.created_by)

    lines = list(
        db.execute(
            select(QuoteResultLine).where(QuoteResultLine.quote_id == quote_id).order_by(QuoteResultLine.id)
        ).scalars().all()
    )
    sku_ids = {ln.sku_id for ln in lines}
    skus = (
        {s.id: s for s in db.execute(select(SKU).where(SKU.id.in_(sku_ids))).scalars().all()}
        if sku_ids else {}
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "КП"

    bold = Font(bold=True)
    ws.append(["Дата", date.today().isoformat()])
    ws.append(["КП №", quote_id])
    ws.append(["Менеджер", manager_name])
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold

    ws.append([])

    headers = ["Код", "Наименование", "Ед. изм.", "Кол-во", "Примечание"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    for ln in lines:
        s = skus.get(ln.sku_id)
        ws.append([
            s.code if s else str(ln.sku_id),
            s.name if s else "",
            s.unit if s else "",
            ln.qty,
            ln.note or "",
        ])

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 30

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
